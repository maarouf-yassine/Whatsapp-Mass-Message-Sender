# -*- coding: utf-8 -*-
"""
Created on Sat Aug 22 20:28:27 2020

@author: maarouf
"""
#################

#Make Sure You Have Whatsapp Web INSTALLED on your device 
#Make Sure your phone numbers are stored on and excel sheet
#Check The excel sheet example in the Repo to see the format (make sure to include the country code without a '+' sign )
#You May Need To Install The Libraries Below
#Copy The text that you want to share through whatsapp BEFORE running this script 
#When you run the script, don't use the mouse and keyboard till it ends 

#################
import time
import pyperclip
from keyboard import press
import pynput
from pynput.keyboard import Key, Controller
keyboard = Controller()
#read the excel sheet 
from openpyxl import load_workbook
wb = load_workbook(r'location of file on computer')  # Work Book
ws = wb.get_sheet_by_name('Sheet1')  # Work Sheet , you can choose any sheet in case you have multiple ones 
column = ws['A']  # Column , A used to get the very first column where the numbers are present
column_list = [str(column[x].value) for x in range(len(column))]
print(column_list)

import webbrowser

url = 'https://api.whatsapp.com/send?phone=' #url that opens whatsapp web application on your device 
webbrowser.register('firefox', #or use any other web browser (chrome,...)
	None,
	webbrowser.BackgroundBrowser("C://Program Files//Mozilla Firefox//firefox.exe")) #location of the broser on your device 

for x in column_list: #looping through every number in the list 
    webbrowser.get('firefox').open(url+x) #url opened on browser become https://api.whatsapp.com/send?phone=xxxxxxxxxxxx where the phone number is included in the end 
#whatsapp web application opens on your device and we wait for 10 seconds to make sure that it started a chat with the non-saved phone number 
    time.sleep(10)
#this script pastes the copied text and presses enter to send the message 
    keyboard.press(Key.ctrl)
    keyboard.press('v')
    keyboard.release('v')
    keyboard.release(Key.ctrl)
    press('enter')
#wait till this script ends and don't use the mouse/keyboard while running 
